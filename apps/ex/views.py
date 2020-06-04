from django.shortcuts import render, redirect, reverse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, StreamingHttpResponse
# from django.db.models import Count,Sum
from django.db.models.aggregates import Aggregate, Count, Sum
# Create your views here.
from django.utils import timezone
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from plans.forms import PlanAddInfoForm
from .models import ScheduleInifo, ProblemInfo, ScheduleFileInifo, ProblemFileInifo, FeedbackFileInifo, ReportFileInifo
from .models import RectificationPrjectInfo, RectificationProblemInfo, RectificationFeedbackInfo, \
    RectificationFeedbackFileInfo
from plans.models import PlansInifo, AuditMemberInfo, AuditCompanyInfo, AllNoticeInfo
from users.models import UserProfile
from django.db.models import Q
import os
import json
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, logout, login
import datetime
from django.core.files import File
import pandas as pd
from openpyxl import Workbook, load_workbook
# import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from users.tests import coding
from openpyxl.styles import Color, Font, Alignment, PatternFill, Border, Side, Protection, colors
from .tools import num2column


@login_required(login_url='users/user_login/')
def schedule_list(request, project_id):
    keyword = request.GET.get('keyword', '')

    sc_list = ScheduleInifo.objects.filter(project_name_id=int(project_id), status=True)
    user_list = sc_list.filter(username_id=request.user.id).order_by('-add_time')
    member_list = sc_list.exclude(username_id=request.user.id).order_by('-add_time')
    if keyword:
        user_list = sc_list.filter(username_id=request.user.id).order_by('-add_time').filter(
            Q(schedule_title__icontains=keyword) | Q(schedule_content__icontains=keyword) | Q(
                schedule_feedback__icontains=keyword) | Q(audit_company__audit_company__icontains=keyword) | Q(
                audit_company__audit_person__icontains=keyword))
        member_list = sc_list.exclude(username_id=request.user.id).order_by('-add_time').filter(
            Q(schedule_title__icontains=keyword) | Q(schedule_content__icontains=keyword) | Q(
                schedule_feedback__icontains=keyword) | Q(audit_company__audit_company__icontains=keyword) | Q(
                audit_company__audit_person__icontains=keyword))
    title = PlansInifo.objects.filter(id=int(project_id), status=True).first()
    this_date = datetime.date.today()
    member_name_list = AuditMemberInfo.objects.filter(project_name_id=int(project_id))
    member_user = member_name_list.filter(username_id=request.user.id, status=True).first()
    sc_file_list = ScheduleFileInifo.objects.filter(status=True, project_name_id=int(project_id))
    audit_company = AuditCompanyInfo.objects.filter(status=True, project_name_id=int(project_id))
    switch = 'close'
    if title.project_status == 'ex' and member_user:
        switch = 'open'

    return render(request, 'ex/schedule_list.html', {
        'project_id': project_id,
        'sc_list': sc_list,
        'user_list': user_list,
        'member_list': member_list,
        'title': title,
        'this_date': this_date,
        'member_name': member_name_list,
        'sc_file_list': sc_file_list,
        'audit_company': audit_company,
        'switch': switch,
        'member_user': member_user

    })


def add_schedule(request):
    if request.method == "POST":
        audit_company = request.POST.get('audit_company')
        schedule_title = request.POST.get('schedule_title', '')
        schedule_content = request.POST.get('schedule_content', '')
        schedule_feedback = request.POST.get('schedule_feedback', '')
        start_time = request.POST.get('start_time', '')
        end_time = request.POST.get('end_time', '')
        project_id = request.POST.get('project_id')
        # return HttpResponse('开发中')
        try:
            a = ScheduleInifo()
            a.company_id_id = request.user.company_name_id
            a.project_name_id = int(project_id)
            a.username_id = request.user.id
            if int(audit_company) > 0:
                a.audit_company_id = int(audit_company)
            a.schedule_title = schedule_title
            a.schedule_content = schedule_content
            a.schedule_feedback = schedule_feedback
            a.start_time = datetime.datetime.strptime(start_time, '%Y-%m-%d')
            a.end_time = datetime.datetime.strptime(end_time, '%Y-%m-%d')
            a.save()
            return JsonResponse({'status': 'ok', 'msg': '提交成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '提交失败'})


def schedule_del(request):
    schedule_id = request.GET.get('schedule_id')
    project_id = request.GET.get('project_id')
    schedule_list = ScheduleInifo.objects.filter(status=True, id=int(schedule_id), project_name_id=int(project_id))
    schedule = schedule_list[0]
    try:
        a = schedule
        a.status = False
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '删除失败'})


def schedule_file(request):
    schedule_file1 = request.FILES.get('schedule_file')
    project_id = request.POST.get('project_id')
    audit_company = request.POST.get('audit_company')
    try:
        a = ScheduleFileInifo()
        a.project_name_id = int(project_id)
        a.company_id_id = request.user.company_name_id
        if int(audit_company) > 0:
            a.audit_company_id = int(audit_company)
        a.username_id = request.user.id
        a.schedule_file = schedule_file1
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '附件上传成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '附件上传失败'})


def schedule_file_del(request):
    schedule_file_id = request.GET.get('schedule_file_id')
    project_id = request.GET.get('project_id')
    schedule_file_list1 = ScheduleFileInifo.objects.filter(status=True, id=int(schedule_file_id))
    schedule_file1 = schedule_file_list1[0]
    try:
        a = schedule_file1
        a.status = False
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '删除失败'})


@login_required(login_url='users/user_login/')
def schedule_detail(request, schedule_id):
    schedule_info = ScheduleInifo.objects.filter(id=int(schedule_id))
    user = schedule_info.first()
    project_id = user.project_name.id
    audit_company = AuditCompanyInfo.objects.filter(status=True, project_name_id=project_id)
    sc_file_list = ScheduleFileInifo.objects.filter(status=True, project_name=project_id)
    member_name_list = AuditMemberInfo.objects.filter(project_name_id=int(project_id))
    member_user = member_name_list.filter(username_id=request.user.id).first()
    title = PlansInifo.objects.filter(id=int(project_id), status=True).first()
    switch = 'close'
    if title.project_status == 'ex' and member_user:
        switch = 'open'
    return render(request, 'ex/schedule_detail.html', {
        'schedule_id': schedule_id,
        'user': user,
        'project_id': project_id,
        'audit_company': audit_company,
        'sc_file_list': sc_file_list,
        'member_name': member_name_list,
        'member_user': member_user,
        'title': title,
        'switch': switch
    })


def fix_schedule(request):
    if request.method == "POST":
        audit_company = request.POST.get('audit_company')
        schedule_title = request.POST.get('schedule_title', '')
        schedule_content = request.POST.get('schedule_content', '')
        schedule_feedback = request.POST.get('schedule_feedback', '')
        start_time = request.POST.get('start_time', '')
        end_time = request.POST.get('end_time', '')
        project_id = request.POST.get('project_id')
        schedule_id = request.POST.get('schedule_id')
        # return HttpResponse('开发中')
        try:
            a1 = ScheduleInifo.objects.filter(id=int(schedule_id))
            a = a1[0]
            if int(audit_company) > 0:
                a.audit_company_id = int(audit_company)
            a.schedule_title = schedule_title
            a.schedule_content = schedule_content
            a.schedule_feedback = schedule_feedback
            a.start_time = datetime.datetime.strptime(start_time, '%Y-%m-%d')
            a.end_time = datetime.datetime.strptime(end_time, '%Y-%m-%d')
            a.save()
            return JsonResponse({'status': 'ok', 'msg': '提交成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '提交失败'})


@login_required(login_url='users/user_login/')
def problem_list(request, project_id):
    do_problem_list = ProblemInfo.objects.filter(status=True, project_name_id=int(project_id),
                                                 problem_status='do').order_by('-add_time')
    co_sg_problem_list = ProblemInfo.objects.filter(status=True, project_name_id=int(project_id), problem_status='co',
                                                    problem_type='sg').order_by('-add_time')
    co_re_problem_list = ProblemInfo.objects.filter(status=True, project_name_id=int(project_id), problem_status='co',
                                                    problem_type='re').order_by('-add_time')
    el_problem_list = ProblemInfo.objects.filter(status=True, project_name_id=int(project_id),
                                                 problem_status='el').order_by('-add_time')
    problem_file_list = ProblemFileInifo.objects.filter(status=True, project_name_id=int(project_id)).order_by(
        '-add_time')
    user = AuditMemberInfo.objects.filter(status=True, project_name_id=project_id, username_id=request.user.id).first()
    switch = 'colse'
    project_is = PlansInifo.objects.filter(status=True, id=int(project_id), project_status='ex').first()
    if project_is and user:
        switch = 'open'

    return render(request, 'ex/problem_list.html', {
        'project_id': project_id,
        'do_problem_list': do_problem_list,
        'co_sg_problem_list': co_sg_problem_list,
        'co_re_problem_list': co_re_problem_list,
        'el_problem_list': el_problem_list,
        'problem_file_list': problem_file_list,
        'switch': switch
    })


@login_required(login_url='users/user_login/')
def problem_detail(request, problem_id):
    problem_info = ProblemInfo.objects.filter(status=True, id=int(problem_id)).first()
    project_id = problem_info.project_name_id
    user = AuditMemberInfo.objects.filter(status=True, project_name_id=project_id, username_id=request.user.id).first()
    switch = 'colse'
    project_is = PlansInifo.objects.filter(status=True, id=int(project_id), project_status='ex').first()
    if project_is and user:
        switch = 'open'

    audit_company = AuditCompanyInfo.objects.filter(status=True, project_name_id=int(project_id))
    problem_file_list = ProblemFileInifo.objects.filter(status=True, problem_id=int(problem_id))
    return render(request, 'ex/problem_detail.html', {
        'project_id': project_id,
        'problem_info': problem_info,
        'user': user,
        'switch': switch,
        'audit_company': audit_company,
        'problem_file_list': problem_file_list

    })


def problem_add(request, project_id):
    audit_company = AuditCompanyInfo.objects.filter(project_name_id=int(project_id), status=True)
    return render(request, 'ex/problem_add.html', {
        'project_id': project_id,
        'audit_company': audit_company
    })


def auto_problem_class(request):
    problem_class = request.GET.get('problem_class')
    project_id = request.GET.get('project_id')
    if problem_class:
        project_info_list = PlansInifo.objects.filter(status=True, id=int(project_id))
        project_info = project_info_list[0]
        result_list = ProblemInfo.objects.filter(company_id_id=project_info.company_id_id,
                                                 problem_class__icontains=problem_class, status=True)
        result = list(set(result1.problem_class for result1 in result_list))
        if len(result) == 1:
            result = result[0]
            return JsonResponse({'status': 'ok', 'msg': result})
        elif result_list:
            result = '1' + str(result)
            return JsonResponse({'status': 'false', 'msg': result})
        else:
            pass


def add_problem(request):
    if request.method == "POST":
        project_id = request.POST.get("project_id")
        audit_company = request.POST.get('audit_company')
        problem_title = request.POST.get('problem_title', '')
        problem_content = request.POST.get('problem_content', '')
        problem_feedback = request.POST.get('problem_feedback', '')
        problem_class = request.POST.get('problem_class', '')
        problem_type = request.POST.get('problem_type', '')
        problem_status = request.POST.get('problem_status', '')
        try:
            a = ProblemInfo()
            a.company_id_id = request.user.company_name.id
            a.username_id = request.user.id
            if int(audit_company) > 0:
                a.audit_company_id = int(audit_company)
            a.project_name_id = int(project_id)
            a.problem_title = problem_title
            a.problem_content = problem_content
            a.problem_feedback = problem_feedback
            a.problem_class = problem_class
            a.problem_type = problem_type
            a.problem_status = problem_status
            a.save()
            # AllNoticeFileInfo.objects.create(company_id=int(company_id),notice_name=int(notice_id),notice_file=notice_file1)
            return JsonResponse({'status': 'ok', 'msg': '新增成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '更新失败'})


def fix_problem(request):
    if request.method == "POST":
        problem_id = request.POST.get('problem_id')
        project_id = request.POST.get("project_id")
        audit_company = request.POST.get('audit_company')
        problem_title = request.POST.get('problem_title', '')
        problem_content = request.POST.get('problem_content', '')
        problem_feedback = request.POST.get('problem_feedback', '')
        problem_class = request.POST.get('problem_class', '')
        problem_type = request.POST.get('problem_type', '')
        problem_status = request.POST.get('problem_status', '')
        try:
            a1 = ProblemInfo.objects.filter(status=True, id=int(problem_id))
            a = a1[0]
            a.company_id_id = request.user.company_name.id
            if int(audit_company) > 0:
                a.audit_company_id = int(audit_company)
            a.project_name_id = int(project_id)
            a.problem_title = problem_title
            a.problem_content = problem_content
            a.problem_feedback = problem_feedback
            a.problem_class = problem_class
            a.problem_type = problem_type
            a.problem_status = problem_status
            a.save()
            # AllNoticeFileInfo.objects.create(company_id=int(company_id),notice_name=int(notice_id),notice_file=notice_file1)
            return JsonResponse({'status': 'ok', 'msg': '修改成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '修改失败'})


def add_problem_file(request):
    if request.method == "POST":
        project_id = request.POST.get('project_id')
        problem_file = request.FILES.get('problem_file')
        problem_id = request.POST.get('problem_id')
        audit_company = request.POST.get('audit_company')
        try:
            a = ProblemFileInifo()
            a.project_name_id = int(project_id)
            a.company_id_id = request.user.company_name_id
            if int(audit_company) > 0:
                a.audit_company_id = int(audit_company)
            a.username_id = request.user.id
            a.problem_file = problem_file
            a.problem_id_id = int(problem_id)
            a.save()
            return JsonResponse({'status': 'ok', 'msg': '附件上传成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '附件上传失败'})


def del_problem_file(request):
    problem_file_id = request.GET.get('problem_file')
    problem_id = request.GET.get('problem_id')
    problem_file_list1 = ProblemFileInifo.objects.filter(status=True, id=int(problem_file_id))
    problem_file1 = problem_file_list1[0]
    try:
        a = problem_file1
        a.status = False
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '删除失败'})


@login_required(login_url='users/user_login/')
def report_detail(request, project_id):
    feedback_file_list = FeedbackFileInifo.objects.filter(status=True, project_name_id=int(project_id)).order_by(
        'add_time')
    switch = 'close'
    project_info = PlansInifo.objects.filter(status=True, id=int(project_id), project_status='ex').first()
    if project_info:
        switch = 'open'
    member = AuditMemberInfo.objects.filter(status=True, username_id=request.user.id,
                                            project_name_id=int(project_id)).first()
    notice_info = AllNoticeInfo.objects.filter(status=True, company_id_id=request.user.company_name_id,
                                               project_name_id=int(project_id), notice_type='report').first()
    report_type1 = ['fi', 'fifi']
    report_fi_file = ReportFileInifo.objects.filter(status=True, project_name_id=int(project_id),
                                                    report_type__in=report_type1).order_by('add_time')
    report_type2 = ['pr', 'prfi']
    report_pr_file = ReportFileInifo.objects.filter(status=True, project_name_id=int(project_id),
                                                    report_type__in=report_type2).order_by('add_time')

    return render(request, 'ex/report_detail.html', {
        'project_id': project_id,
        'feedback_file_list': feedback_file_list,
        'switch': switch,
        'member': member,
        'notice_info': notice_info,
        'report_fi_file': report_fi_file,
        'report_pr_file': report_pr_file
    })


def add_feedback_file(request):
    if request.method == "POST":
        project_id = request.POST.get('project_id')
        feedback_file = request.FILES.get('feedback_file')
        feedback_type = request.POST.get('feedback_type')
        try:
            a = FeedbackFileInifo()
            a.project_name_id = int(project_id)
            a.company_id_id = request.user.company_name_id
            a.username_id = request.user.id
            a.feedback_file = feedback_file
            a.feedback_type = feedback_type
            a.save()
            return JsonResponse({'status': 'ok', 'msg': '附件上传成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '附件上传失败'})


def del_feedback_file(request):
    feedback_file_id = request.GET.get('feedback_file')
    feedback_file_list1 = FeedbackFileInifo.objects.filter(status=True, id=int(feedback_file_id))
    feedback_file1 = feedback_file_list1[0]
    try:
        a = feedback_file1
        a.status = False
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '删除失败'})


def notice_info(request):
    if request.method == "POST":
        project_id = request.POST.get('project_id')
        notice_name = request.POST.get('notice_name')
        notice_NO = request.POST.get('notice_NO')
        notice_info = AllNoticeInfo.objects.filter(status=True, company_id=int(request.user.company_name_id),
                                                   project_name=int(project_id), notice_type='report')
        if notice_info:
            notice_info1 = notice_info[0]
            try:
                a = notice_info1
                a.notice_name = notice_name
                a.notice_NO = notice_NO
                a.notice_type = 'report'
                a.project_name_id = int(project_id)
                a.company_id_id = int(request.user.company_name_id)
                a.save()
                return JsonResponse({'status': 'ok', 'msg': '修改成功1'})
            except Exception as e:
                return JsonResponse({'status': 'false', 'msg': '修改失败1'})

        else:
            notice_info = AllNoticeInfo()
            try:
                a = notice_info
                a.notice_name = notice_name
                a.notice_NO = notice_NO
                a.notice_type = 'report'
                a.project_name_id = int(project_id)
                a.company_id_id = int(request.user.company_name_id)
                a.save()
                return JsonResponse({'status': 'ok', 'msg': '修改成功2'})
            except Exception as e:
                return JsonResponse({'status': 'false', 'msg': '修改失败2'})


def add_report_file(request):
    if request.method == "POST":
        project_id = request.POST.get('project_id')
        report_file = request.FILES.get('report_file')
        report_type = request.POST.get('report_type')
        try:
            a = ReportFileInifo()
            a.project_name_id = int(project_id)
            a.company_id_id = request.user.company_name_id
            a.username_id = request.user.id
            a.report_file = report_file
            a.report_type = report_type
            a.save()
            return JsonResponse({'status': 'ok', 'msg': '附件上传成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '附件上传失败'})


def del_report_file(request):
    report_file_id = request.GET.get('report_file')
    report_file_list1 = ReportFileInifo.objects.filter(status=True, id=int(report_file_id))
    report_file1 = report_file_list1[0]
    try:
        a = report_file1
        a.status = False
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '删除失败'})


def clean_doubts(request):
    project_id = request.POST.get('project_id')
    problem_list = ProblemInfo.objects.filter(status=True, project_name_id=int(project_id), problem_status='do')
    if problem_list:
        try:
            for problem in problem_list:
                a = problem
                a.problem_status = 'co'
                a.save()
            return JsonResponse({'status': 'ok', 'msg': '更新成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '更新失败'})
    else:
        return JsonResponse({'status': 'false', 'msg': '没有疑点需要确认'})


def off_project(request):
    if request.method == "POST":
        project_id = request.POST.get('project_id')
        problem_list = ProblemInfo.objects.filter(status=True, project_name_id=int(project_id), problem_status='do')
        if problem_list:
            return JsonResponse({'status': 'false', 'msg': '请确认所有的疑点后，关闭项目'})
        else:
            project_info = PlansInifo.objects.filter(status=True, id=int(project_id)).first()
            problem_list1 = ProblemInfo.objects.filter(status=True, project_name_id=int(project_id),
                                                       problem_status='co', problem_type='re')
            report_NO = AllNoticeInfo.objects.filter(status=True, project_name_id=int(project_id),
                                                     notice_type='report').first()
            if problem_list1:
                try:
                    a = RectificationPrjectInfo()
                    a.project_name_id = project_info.id
                    a.re_status = 'pr'
                    a.company_id_id = request.user.company_name_id
                    a.report_NO_id = report_NO.id
                    a.save()
                    re_project = RectificationPrjectInfo.objects.filter(status=True,
                                                                        project_name_id=project_info.id).first()
                    for problem_info in problem_list1:
                        b = RectificationProblemInfo()
                        b.re_project_id = re_project.id
                        b.re_problem_id = problem_info.id
                        b.company_id_id = request.user.company_name_id
                        b.save()
                    project_old = PlansInifo.objects.filter(status=True, id=int(project_id))
                    c = project_old[0]
                    c.project_status = 'cp'
                    c.save()
                    return JsonResponse({'status': 'ok', 'msg': '关闭成功'})
                except Exception as e:
                    return JsonResponse({'status': 'false', 'msg': '关闭失败'})
            else:
                try:
                    project_old = PlansInifo.objects.filter(status=True, id=int(project_id))
                    c = project_old[0]
                    c.project_status = 'cp'
                    c.save()
                    return JsonResponse({'status': 'ok', 'msg': '关闭成功'})
                except Exception as e:
                    return JsonResponse({'status': 'false', 'msg': '关闭失败'})


@login_required(login_url='users/user_login/')
def re_project(request):
    re_pr_project = RectificationPrjectInfo.objects.filter(status=True, company_id_id=request.user.company_name_id,
                                                           re_status='pr')
    re_fi_project = RectificationPrjectInfo.objects.filter(status=True, company_id_id=request.user.company_name_id,
                                                           re_status='co').order_by('-re_year', 'add_time')
    return render(request, 'ex/re_project.html', {
        're_pr_project': re_pr_project,
        're_fi_project': re_fi_project
    })


@login_required(login_url='users/user_login/')
def re_problem(request, re_project_id):
    re_project = RectificationPrjectInfo.objects.filter(status=True, company_id_id=request.user.company_name_id,
                                                        id=int(re_project_id)).first()
    re_pr_problem_list = RectificationProblemInfo.objects.filter(status=True,
                                                                 company_id_id=request.user.company_name_id,
                                                                 re_project_id=int(re_project_id), re_situation='pr')
    re_co_problem_list = RectificationProblemInfo.objects.filter(status=True,
                                                                 company_id_id=request.user.company_name_id,
                                                                 re_project_id=int(re_project_id), re_situation='co')
    re_un_problem_list = RectificationProblemInfo.objects.filter(status=True,
                                                                 company_id_id=request.user.company_name_id,
                                                                 re_project_id=int(re_project_id), re_situation='un')
    switch = 'close'
    if re_project.re_status == 'pr':
        switch = 'open'
    re_problem_list = RectificationProblemInfo.objects.filter(status=True,
                                                              company_id_id=request.user.company_name_id,
                                                              re_project_id=int(re_project_id)).order_by('re_situation')
    return render(request, 'ex/re_problem.html', {
        're_project_id': re_project_id,
        're_project': re_project,
        're_pr_problem_list': re_pr_problem_list,
        're_co_problem_list': re_co_problem_list,
        're_un_problem_list': re_un_problem_list,
        'switch': switch,
        're_problem_list': re_problem_list

    })


def re_finish_project(request):
    re_project_id = request.GET.get('re_project_id')
    re_project_info = RectificationPrjectInfo.objects.filter(status=True, id=int(re_project_id)).first()
    re_pr_problem_list = RectificationProblemInfo.objects.filter(status=True,
                                                                 company_id_id=request.user.company_name_id,
                                                                 re_project_id=int(re_project_id), re_situation='pr')
    if re_pr_problem_list:
        return JsonResponse({'status': 'false', 'msg': '仍有问题在整改中'})
    else:
        try:
            a = re_project_info
            a.re_status = 'co'
            a.save()
            return JsonResponse({'status': 'ok', 'msg': '项目结束成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '项目结束失败'})


@login_required(login_url='users/user_login/')
def re_feedback(request, re_problem_id):
    re_problem = RectificationProblemInfo.objects.filter(status=True, id=int(re_problem_id),
                                                         company_id_id=request.user.company_name_id).first()
    re_project = RectificationPrjectInfo.objects.filter(status=True, company_id_id=request.user.company_name_id,
                                                        id=re_problem.re_project_id).first()
    re_project_id = re_project.id
    re_feedback_list = RectificationFeedbackInfo.objects.filter(status=True, company_id=request.user.company_name_id,
                                                                re_project_name_id=int(re_problem_id)).order_by(
        '-add_time')
    today = datetime.datetime.now().today()
    switch = 'close'
    if re_project.re_status == 'pr':
        switch = 'open'

    re_feedback_nm = list(set(re_feedback.id for re_feedback in re_feedback_list))
    re_feedback_file_list = RectificationFeedbackFileInfo.objects.filter(status=True, re_feedback_id__in=re_feedback_nm)
    return render(request, 'ex/re_feedback.html', {
        're_project': re_project,
        're_project_id': re_project_id,
        're_problem': re_problem,
        're_feedback_list': re_feedback_list,
        'today': today,
        're_problem_id': re_problem_id,
        'switch': switch,
        're_feedback_file_list': re_feedback_file_list
    })


def add_feedback(request):
    if request.method == "POST":
        re_department = request.POST.get('re_department')
        re_person = request.POST.get('re_person')
        re_improve = request.POST.get('re_improve')
        re_date = request.POST.get('re_date')
        re_situation = request.POST.get('re_situation')
        accountability_is = request.POST.get('accountability_is')
        accountability_content = request.POST.get('accountability_content')
        announcement_is = request.POST.get('announcement_is')
        remarks = request.POST.get('remarks')
        re_problem_id = request.POST.get('re_problem_id')
        try:
            a = RectificationFeedbackInfo()
            a.re_project_name_id = int(re_problem_id)
            a.company_id_id = request.user.company_name_id
            a.re_improve = re_improve
            a.re_department = re_department
            a.re_person = re_person
            a.re_date = datetime.datetime.strptime(re_date, '%Y-%m-%d')
            a.re_situation = re_situation
            a.accountability_is = accountability_is
            a.accountability_content = accountability_content
            a.announcement_is = announcement_is
            a.remarks = remarks
            a.username_id = request.user.id
            a.save()
            re_problem = RectificationProblemInfo.objects.filter(status=True, id=int(re_problem_id))
            b = re_problem[0]
            b.re_improve = re_improve
            b.re_department = re_department
            b.re_person = re_person
            b.re_date = datetime.datetime.strptime(re_date, '%Y-%m-%d')
            b.re_situation = re_situation
            b.accountability_is = accountability_is
            b.accountability_content = accountability_content
            b.announcement_is = announcement_is
            b.remarks = remarks
            b.username_id = request.user.id
            b.save()
            return JsonResponse({'status': 'ok', 'msg': '添加成功'})
        except Exception as e:
            return JsonResponse({'status': 'false', 'msg': '添加失败'})


def del_feedback(request):
    feedback_id = request.GET.get('feedback_id')
    feedback_list1 = RectificationFeedbackInfo.objects.filter(status=True, id=int(feedback_id))
    feedback_info = feedback_list1[0]
    try:
        problem_info_id = feedback_info.re_project_name_id
        a = feedback_info
        a.status = False
        a.save()
        feedback_info2 = RectificationFeedbackInfo.objects.filter(status=True,
                                                                  re_project_name_id=problem_info_id).order_by(
            '-add_time').first()
        problem_info = RectificationProblemInfo.objects.filter(status=True, id=problem_info_id).first()
        if feedback_info2:
            b = problem_info
            b.re_improve = feedback_info2.re_improve
            b.re_department = feedback_info2.re_department
            b.re_person = feedback_info2.re_person
            b.re_date = feedback_info2.re_date
            b.re_situation = feedback_info2.re_situation
            b.accountability_is = feedback_info2.accountability_is
            b.accountability_content = feedback_info2.accountability_content
            b.announcement_is = feedback_info2.announcement_is
            b.remarks = feedback_info2.remarks
            b.username_id = feedback_info2.username_id
            b.save()
        else:
            b = problem_info
            b.re_improve = None
            b.re_department = None
            b.re_person = None
            b.re_date = datetime.datetime.today()
            b.re_situation = 'pr'
            b.accountability_is = 'n'
            b.accountability_content = None
            b.announcement_is = 'n'
            b.remarks = None
            b.username_id = None
            b.save()
        return JsonResponse({'status': 'ok', 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '删除失败'})


def re_feedback_file(request):
    re_feedback_id = request.POST.get('re_feedback_id')
    re_feedback_file = request.FILES.get('re_feedback_file')
    try:
        a = RectificationFeedbackFileInfo()
        a.company_id_id = request.user.company_name_id
        a.username_id = request.user.id
        a.re_feedback_id = int(re_feedback_id)
        a.re_file = re_feedback_file
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '新增成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '新增失败'})


def re_feedback_file_del(request):
    re_feedback_file_id = request.GET.get('re_feedback_file_id')
    re_feedback_file_list1 = RectificationFeedbackFileInfo.objects.filter(status=True, id=int(re_feedback_file_id))
    re_feedback_file = re_feedback_file_list1[0]
    try:
        a = re_feedback_file
        a.status = False
        a.save()
        return JsonResponse({'status': 'ok', 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'status': 'false', 'msg': '删除失败'})


def re_problem_exportxl(request, re_project_id):
    # re_project_id=request.GET.get('re_project_id')
    re_project = RectificationPrjectInfo.objects.filter(status=True, company_id_id=request.user.company_name_id,
                                                        id=int(re_project_id)).first()
    re_problem_list = RectificationProblemInfo.objects.filter(status=True,
                                                              company_id_id=int(request.user.company_name_id),
                                                              re_project_id=int(re_project_id))
    workbook = Workbook()
    data_re_problem = workbook.create_sheet(u"编码", 0)
    border = Border(left=Side(style='thick', color=colors.RED),
                    right=Side(style='medium', color=colors.BLACK),
                    top=Side(style='double', color=colors.BLUE),
                    bottom=Side(style='dotted', color=colors.BLACK))
    # workbook=workbook.active
    re_problem_header = ["序号", "问题名称", "整改状态", "问题内容", "被审计公司", "被审计人", "整改问题编码"]
    data_re_problem.append(re_problem_header)
    # data_re_problem.cell(row=1,column=1,value="序号")
    # data_re_problem.cell(row=1, column=2, value="名称")
    # data_re_problem["{}{}".format(num2column(1), 1)].border = border
    # data_re_problem["{}{}".format(num2column(2), 1)].border = border
    # data_re_problem["{}{}".format(num2column(3), 1)].border = border
    # data_re_problem["{}{}".format(num2column(4), 1)].border = border
    # data_re_problem["{}{}".format(num2column(5), 1)].border = border
    # data_re_problem["{}{}".format(num2column(6), 1)].border = border
    # data_re_problem["{}{}".format(num2column(7), 1)].border = border
    num = 0
    if re_problem_list:
        for re_problem in re_problem_list:
            data_re_problem.cell(row=num + 2, column=1).value = num + 1
            # data_re_problem["{}{}".format(num2column(1),num+2)].border=border
            data_re_problem.cell(row=num + 2, column=2).value = re_problem.re_problem.problem_title
            # data_re_problem["{}{}".format(num2column(2), num + 2)].border = border
            if re_problem.re_situation == 'pr':
                data_re_problem.cell(row=num + 2, column=3).value = '整改中'
                # data_re_problem["{}{}".format(num2column(num + 2), 3)].border = border
            elif re_problem.re_situation == 'co':
                data_re_problem.cell(row=num + 2, column=3).value = '整改完成'
                # data_re_problem["{}{}".format(num2column(num + 2), 3)].border = border
            else:
                data_re_problem.cell(row=num + 2, column=3).value = '无法整改'
                # data_re_problem["{}{}".format(num2column(num + 2), 3)].border = border
            data_re_problem.cell(row=num + 2, column=4).value = re_problem.re_problem.problem_content
            # data_re_problem["{}{}".format(num2column(num + 2), 4)].border = border
            data_re_problem.cell(row=num + 2, column=5).value = re_problem.re_problem.audit_company.audit_company
            # data_re_problem["{}{}".format(num2column(num + 2), 5)].border = border
            data_re_problem.cell(row=num + 2, column=6).value = re_problem.re_problem.audit_company.audit_person
            # data_re_problem["{}{}".format(num2column(num + 2), 6)].border = border
            data_re_problem.cell(row=num + 2, column=7).value = coding(re_problem.id)
            # data_re_problem["{}{}".format(num2column(num + 2), 7)].border = border
            num += 1
    #         # aa['问题名称']=re_problem.re_problem.problem_title
    #         # aa['整改状态']=re_problem.re_situation
    #         # aa['问题内容']=re_problem.re_problem.problem_content
    #         # aa['被审计公司']=re_problem.re_problem.audit_company.audit_company
    #         # aa['被审计人']=re_problem.re_problem.audit_company.audit_person
    #         # aa['整改问题编码']=re_problem.coding
    #         # re_problem_list2.insert(num,aa)

    file_name = re_project.project_name.project_name + '--问题编码' + datetime.datetime.today().strftime(
        '%Y-%m-%d %H:%M:%S')
    wb = save_virtual_workbook(workbook)
    # response = StreamingHttpResponse(streaming_content=wb,content_type='application/octet-stream')
    response = HttpResponse(content=wb, content_type='application/octet-stream')
    # response["content_type"] = 'application/octet-stream'
    # response['Content-Type'] = 'application/vnd.ms-excel'
    # response['Content-Type']= 'application/x-xlsx'

    # response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % "活动数据"
    response["Content-Disposition"] = 'attachment; filename={}.xls'.format(file_name.encode().decode('latin-1'))
    # workbook.close()
    # response = HttpResponse(save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename=%s.xls' % file_name
    return response
    # aa={}
    # re_problem_list2=[]
    # num=0

    # pf=pd.DataFrame(list(re_problem_list2))
    # columns_map={
    #
    # }
    # # pf.rename(columns=columns_map, inplace=True)
    # # # 将空的单元格替换为空字符
    # # pf.fillna('', inplace=True)
    # try:
    #     filename = 're_problem_coding.xlsx'
    #     BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    #     MEDIA_ROOT = os.path.join(BASE_DIR, 'media')
    #     # root_path = MEDIA_ROOT + '/fileOutput/'
    #     file_path = os.path.join(MEDIA_ROOT,'fileOutput', filename)
    #     # writer=pd.ExcelWriter(file_path)
    #     pf.to_excel(file_path, sheet_name='sheet1', index=False)
    #     # writer.save()
    #     # writer.close()
    #     download_path=os.path.join('fileOutput', filename)
    #     return JsonResponse({'status': 'ok', 'msg': '新增成功','path':download_path})
    # except Exception as e:
    #     return JsonResponse({'status': 'false', 'msg': '新增失败'})
